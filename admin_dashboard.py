import customtkinter as ctk
from tkinter import messagebox, ttk
import tkinter as tk
from tkinter import filedialog
from mysql.connector import Error
import db_config
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from datetime import datetime, timedelta

class AdminDashboard:
    def __init__(self, parent, main_app):
        self.parent = parent
        self.main_app = main_app
        
        # Create persistent database connection
        self.db_connection = None
        self.get_db_connection()
        
        # Create main container
        self.container = ctk.CTkFrame(parent)
        self.container.pack(fill="both", expand=True)
        
        # Create header
        self.create_header()
        
        # Create navigation
        self.create_navigation()
        
        # Create content area
        self.content_frame = ctk.CTkFrame(self.container)
        self.content_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Show dashboard by default
        self.show_dashboard()
    
    def get_db_connection(self):
        """Get or create database connection - reuse existing if available"""
        if self.db_connection is None or not self.db_connection.is_connected():
            try:
                self.db_connection = db_config.create_connection()
            except Error as e:
                messagebox.showerror("Database Error", f"Failed to connect to database: {e}")
                return None
        return self.db_connection
    
    def execute_query(self, query, params=None, fetch=True, dictionary=False):
        """Execute a database query with proper error handling"""
        connection = self.get_db_connection()
        if not connection:
            return None
        
        try:
            cursor = connection.cursor(dictionary=dictionary)
            cursor.execute(query, params or ())
            
            if fetch:
                result = cursor.fetchall()
                cursor.close()
                return result
            else:
                connection.commit()
                cursor.close()
                return True
        except Error as e:
            messagebox.showerror("Database Error", f"Query failed: {e}")
            return None
    
    def __del__(self):
        """Close database connection when dashboard is destroyed"""
        if self.db_connection and self.db_connection.is_connected():
            self.db_connection.close()
    
    def create_header(self):
        header_frame = ctk.CTkFrame(self.container, height=60)
        header_frame.pack(fill="x", padx=10, pady=(10, 0))
        header_frame.pack_propagate(False)
        
        # Welcome message
        welcome_label = ctk.CTkLabel(
            header_frame,
            text=f"Admin Dashboard - Welcome, {self.main_app.current_user['full_name']}",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        welcome_label.pack(side="left", padx=20)
        
        # Logout button
        logout_button = ctk.CTkButton(
            header_frame,
            text="Logout",
            font=ctk.CTkFont(size=14),
            width=100,
            command=self.logout
        )
        logout_button.pack(side="right", padx=20)
    
    def create_navigation(self):
        nav_frame = ctk.CTkFrame(self.container, height=50)
        nav_frame.pack(fill="x", padx=10, pady=5)
        nav_frame.pack_propagate(False)
        
        # Navigation buttons
        dashboard_btn = ctk.CTkButton(
            nav_frame,
            text="Dashboard",
            font=ctk.CTkFont(size=14),
            width=120,
            command=self.show_dashboard
        )
        dashboard_btn.pack(side="left", padx=10)
        
        users_btn = ctk.CTkButton(
            nav_frame,
            text="User Management",
            font=ctk.CTkFont(size=14),
            width=140,
            command=self.show_users
        )
        users_btn.pack(side="left", padx=10)
        
        suppliers_btn = ctk.CTkButton(
            nav_frame,
            text="Suppliers",
            font=ctk.CTkFont(size=14),
            width=120,
            command=self.show_suppliers
        )
        suppliers_btn.pack(side="left", padx=10)
        
        reports_btn = ctk.CTkButton(
            nav_frame,
            text="Reports",
            font=ctk.CTkFont(size=14),
            width=120,
            command=self.show_reports
        )
        reports_btn.pack(side="left", padx=10)
    
    def clear_content(self):
        for widget in self.content_frame.winfo_children():
            widget.destroy()
    
    def show_dashboard(self):
        self.clear_content()
        
        # Title
        title_label = ctk.CTkLabel(
            self.content_frame,
            text="Admin Dashboard",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(pady=20)
        
        # Stats frame with responsive grid layout
        stats_frame = ctk.CTkFrame(self.content_frame)
        stats_frame.pack(fill="x", padx=20, pady=20)
        # Configure grid columns for responsive layout
        for i in range(5):
            stats_frame.grid_columnconfigure(i, weight=1, uniform="stats")
        
        try:
            connection = self.get_db_connection()
            if connection:
                cursor = connection.cursor()
                
                # Get total users
                cursor.execute("SELECT COUNT(*) FROM users")
                total_users = cursor.fetchone()[0]
                
                # Get pending staff approvals
                cursor.execute("SELECT COUNT(*) FROM users WHERE user_type = 'staff' AND is_approved = FALSE")
                pending_staff = cursor.fetchone()[0]
                
                # Get total products
                cursor.execute("SELECT COUNT(*) FROM products")
                total_products = cursor.fetchone()[0]
                
                # Get total orders
                cursor.execute("SELECT COUNT(*) FROM orders")
                total_orders = cursor.fetchone()[0]
                
                # Get total revenue
                cursor.execute("SELECT SUM(total_amount) FROM orders WHERE status != 'cancelled'")
                total_revenue = cursor.fetchone()[0] or 0
                
                cursor.close()
                
                # Display stats with responsive layout (using grid configured above)
                self.create_stat_card(stats_frame, "Total Users", total_users, "blue").grid(row=0, column=0, padx=10, pady=20, sticky="ew")
                self.create_stat_card(stats_frame, "Pending Staff", pending_staff, "orange").grid(row=0, column=1, padx=10, pady=20, sticky="ew")
                self.create_stat_card(stats_frame, "Total Products", total_products, "green").grid(row=0, column=2, padx=10, pady=20, sticky="ew")
                self.create_stat_card(stats_frame, "Total Orders", total_orders, "purple").grid(row=0, column=3, padx=10, pady=20, sticky="ew")
                self.create_stat_card(stats_frame, f"Revenue\n${total_revenue:.2f}", "", "green").grid(row=0, column=4, padx=10, pady=20, sticky="ew")
                
        except Error as e:
            messagebox.showerror("Error", f"Failed to load dashboard: {e}")
        
        # Quick actions
        actions_frame = ctk.CTkFrame(self.content_frame)
        actions_frame.pack(fill="x", padx=20, pady=20)
        
        ctk.CTkLabel(
            actions_frame,
            text="Quick Actions",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(pady=10)
        
        buttons_frame = ctk.CTkFrame(actions_frame, fg_color="transparent")
        buttons_frame.pack(pady=10)
        
        ctk.CTkButton(
            buttons_frame,
            text="Approve Staff",
            width=200,
            height=40,
            command=self.show_users
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            buttons_frame,
            text="Manage Suppliers",
            width=200,
            height=40,
            command=self.show_suppliers
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            buttons_frame,
            text="View Reports",
            width=200,
            height=40,
            command=self.show_reports
        ).pack(side="left", padx=10)
    
    def create_stat_card(self, parent, title, value, color):
        # Responsive stat card with relative sizing
        card = ctk.CTkFrame(parent)
        card.configure(height=120)
        card.pack_propagate(False)
        
        ctk.CTkLabel(
            card,
            text=title,
            font=ctk.CTkFont(size=14),
            text_color="gray"
        ).pack(pady=(20, 5))
        
        if value != "":
            ctk.CTkLabel(
                card,
                text=str(value),
                font=ctk.CTkFont(size=36, weight="bold"),
                text_color=color
            ).pack(pady=(0, 20))
        
        return card
    
    def show_users(self):
        self.clear_content()
        
        # Title
        title_label = ctk.CTkLabel(
            self.content_frame,
            text="User Management",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(pady=20)
        
        # Filter frame
        filter_frame = ctk.CTkFrame(self.content_frame)
        filter_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(filter_frame, text="Filter by Type:", font=ctk.CTkFont(size=14)).pack(side="left", padx=10)
        
        self.user_type_var = tk.StringVar(value="All")
        type_menu = ctk.CTkOptionMenu(
            filter_frame,
            variable=self.user_type_var,
            values=["All", "customer", "staff", "admin"],
            width=150,
            command=lambda x: self.load_users()
        )
        type_menu.pack(side="left", padx=10)
        
        ctk.CTkLabel(filter_frame, text="Status:", font=ctk.CTkFont(size=14)).pack(side="left", padx=10)
        
        self.approval_var = tk.StringVar(value="All")
        approval_menu = ctk.CTkOptionMenu(
            filter_frame,
            variable=self.approval_var,
            values=["All", "Approved", "Pending"],
            width=150,
            command=lambda x: self.load_users()
        )
        approval_menu.pack(side="left", padx=10)
        
        # Users table with improved styling
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Create Treeview with better styling
        columns = ("ID", "Name", "Email", "Phone", "Type", "Status", "Registered")
        
        # Style configuration for Treeview
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview",
                       background="#2b2b2b",
                       foreground="white",
                       fieldbackground="#2b2b2b",
                       rowheight=30,
                       font=("Segoe UI", 11))
        style.configure("Treeview.Heading",
                       background="#1f538d",
                       foreground="white",
                       font=("Segoe UI", 12, "bold"),
                       padding=(5, 8))
        style.map("Treeview",
                 background=[("selected", "#0066cc")],
                 foreground=[("selected", "white")])
        
        self.users_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15, style="Treeview")
        
        # Configure columns with better widths
        column_widths = {"ID": 60, "Name": 200, "Email": 250, "Phone": 130, "Type": 100, "Status": 100, "Registered": 180}
        for col in columns:
            self.users_tree.heading(col, text=col, anchor="w")
            self.users_tree.column(col, width=column_widths.get(col, 100), anchor="w", minwidth=80)
        
        # Scrollbar with better styling
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.users_tree.yview)
        self.users_tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack with expand for responsiveness
        self.users_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Action buttons
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkButton(
            button_frame,
            text="Approve Selected Staff",
            width=180,
            command=self.approve_staff
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            button_frame,
            text="Delete Selected User",
            width=180,
            fg_color="red",
            hover_color="darkred",
            command=self.delete_user
        ).pack(side="left", padx=10)
        
        self.load_users()
    
    def load_users(self):
        # Clear existing items
        for item in self.users_tree.get_children():
            self.users_tree.delete(item)
        
        try:
            connection = self.get_db_connection()
            if connection:
                cursor = connection.cursor(dictionary=True)
                
                # Build query
                query = "SELECT * FROM users WHERE 1=1"
                params = []
                
                # Type filter
                if self.user_type_var.get() != "All":
                    query += " AND user_type = %s"
                    params.append(self.user_type_var.get())
                
                # Approval filter
                if self.approval_var.get() == "Approved":
                    query += " AND is_approved = TRUE"
                elif self.approval_var.get() == "Pending":
                    query += " AND is_approved = FALSE"
                
                query += " ORDER BY created_at DESC"
                
                cursor.execute(query, params)
                users = cursor.fetchall()
                
                for user in users:
                    status = "Approved" if user['is_approved'] else "Pending"
                    tag = "pending" if not user['is_approved'] and user['user_type'] == 'staff' else ""
                    
                    # Construct full_name from first_name and last_name if needed
                    full_name = user.get('full_name') or f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
                    
                    self.users_tree.insert("", "end", values=(
                        user['id'],
                        full_name,
                        user['email'],
                        user['phone_number'] or 'N/A',
                        user['user_type'].capitalize(),
                        status,
                        user['created_at'].strftime('%Y-%m-%d %H:%M')
                    ), tags=(tag,))
                
                # Highlight pending staff with better colors
                self.users_tree.tag_configure("pending", background="#fff9c4", foreground="#333333")
                
                cursor.close()
                
        except Error as e:
            messagebox.showerror("Error", f"Failed to load users: {e}")
    
    def approve_staff(self):
        selected = self.users_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a staff member to approve")
            return
        
        user_id = self.users_tree.item(selected[0])['values'][0]
        user_type = self.users_tree.item(selected[0])['values'][4]
        status = self.users_tree.item(selected[0])['values'][5]
        
        if user_type != "Staff":
            messagebox.showwarning("Warning", "Selected user is not a staff member")
            return
        
        if status == "Approved":
            messagebox.showinfo("Info", "This staff member is already approved")
            return
        
        try:
            connection = self.get_db_connection()
            if connection:
                cursor = connection.cursor()
                cursor.execute(
                    "UPDATE users SET is_approved = TRUE WHERE id = %s",
                    (user_id,)
                )
                connection.commit()
                cursor.close()
                
                messagebox.showinfo("Success", "Staff member approved successfully")
                self.load_users()
                
        except Error as e:
            messagebox.showerror("Error", f"Failed to approve staff: {e}")
    
    def delete_user(self):
        selected = self.users_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a user to delete")
            return
        
        user_name = self.users_tree.item(selected[0])['values'][1]
        user_type = self.users_tree.item(selected[0])['values'][4]
        
        if user_type == "Admin":
            messagebox.showwarning("Warning", "Cannot delete admin users")
            return
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete user '{user_name}'?"):
            user_id = self.users_tree.item(selected[0])['values'][0]
            
            try:
                connection = self.get_db_connection()
                if connection:
                    cursor = connection.cursor()
                    cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
                    connection.commit()
                    cursor.close()
                    
                    messagebox.showinfo("Success", "User deleted successfully")
                    self.load_users()
                    
            except Error as e:
                messagebox.showerror("Error", f"Failed to delete user: {e}")
    
    def show_suppliers(self):
        self.clear_content()
        
        # Title
        title_label = ctk.CTkLabel(
            self.content_frame,
            text="Supplier Management",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(pady=20)
        
        # Action buttons
        action_frame = ctk.CTkFrame(self.content_frame)
        action_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkButton(
            action_frame,
            text="Add New Supplier",
            width=150,
            command=self.add_supplier
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            action_frame,
            text="Refresh",
            width=100,
            command=self.load_suppliers
        ).pack(side="left", padx=10)
        
        # Suppliers table with improved styling
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Create Treeview with better styling
        columns = ("ID", "Name", "Contact Person", "Email", "Phone", "Address")
        
        # Style configuration for Treeview (same as users table)
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview",
                       background="#2b2b2b",
                       foreground="white",
                       fieldbackground="#2b2b2b",
                       rowheight=30,
                       font=("Segoe UI", 11))
        style.configure("Treeview.Heading",
                       background="#1f538d",
                       foreground="white",
                       font=("Segoe UI", 12, "bold"),
                       padding=(5, 8))
        style.map("Treeview",
                 background=[("selected", "#0066cc")],
                 foreground=[("selected", "white")])
        
        self.suppliers_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15, style="Treeview")
        
        # Configure columns with better widths
        column_widths = {"ID": 60, "Name": 180, "Contact Person": 180, "Email": 220, "Phone": 130, "Address": 250}
        for col in columns:
            self.suppliers_tree.heading(col, text=col, anchor="w")
            self.suppliers_tree.column(col, width=column_widths.get(col, 100), anchor="w", minwidth=80)
        
        # Scrollbar with better styling
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.suppliers_tree.yview)
        self.suppliers_tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack with expand for responsiveness
        self.suppliers_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Action buttons
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkButton(
            button_frame,
            text="Edit Selected",
            width=150,
            command=self.edit_supplier
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            button_frame,
            text="Delete Selected",
            width=150,
            fg_color="red",
            hover_color="darkred",
            command=self.delete_supplier
        ).pack(side="left", padx=10)
        
        self.load_suppliers()
    
    def load_suppliers(self):
        # Check if treeview exists
        if not hasattr(self, 'suppliers_tree') or self.suppliers_tree is None:
            return
        
        # Clear existing items
        for item in self.suppliers_tree.get_children():
            self.suppliers_tree.delete(item)
        
        try:
            # Use fresh connection to get latest data
            connection = db_config.create_connection()
            if connection:
                cursor = connection.cursor(dictionary=True)
                cursor.execute("SELECT * FROM suppliers ORDER BY name")
                suppliers = cursor.fetchall()
                
                for supplier in suppliers:
                    self.suppliers_tree.insert("", "end", values=(
                        supplier['id'],
                        supplier['name'],
                        supplier['contact_person'] or 'N/A',
                        supplier['email'] or 'N/A',
                        supplier['phone'] or 'N/A',
                        supplier['address'] or 'N/A'
                    ))
                
                cursor.close()
                connection.close()
                
        except Error as e:
            messagebox.showerror("Error", f"Failed to load suppliers: {e}")
    
    def add_supplier(self):
        AddSupplierDialog(self.main_app, self)
    
    def edit_supplier(self):
        selected = self.suppliers_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a supplier to edit")
            return
        
        supplier_id = self.suppliers_tree.item(selected[0])['values'][0]
        EditSupplierDialog(self.main_app, self, supplier_id)
    
    def delete_supplier(self):
        selected = self.suppliers_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a supplier to delete")
            return
        
        supplier_name = self.suppliers_tree.item(selected[0])['values'][1]
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete supplier '{supplier_name}'?"):
            supplier_id = self.suppliers_tree.item(selected[0])['values'][0]
            
            try:
                # Use a fresh connection to ensure proper deletion
                connection = db_config.create_connection()
                if connection:
                    cursor = connection.cursor()
                    cursor.execute("DELETE FROM suppliers WHERE id = %s", (supplier_id,))
                    connection.commit()
                    cursor.close()
                    connection.close()
                    
                    messagebox.showinfo("Success", "Supplier deleted successfully")
                    # Refresh supplier list after a brief delay
                    self.main_app.after(100, self.load_suppliers) 
                    
            except Error as e:
                messagebox.showerror("Error", f"Failed to delete supplier: {e}")
    
    def show_reports(self):
        self.clear_content()
        
        # Title and Export button frame
        title_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        title_frame.pack(fill="x", padx=20, pady=20)
        
        title_label = ctk.CTkLabel(
            title_frame,
            text="System Reports",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(side="left")
        
        # Export to Excel button
        export_button = ctk.CTkButton(
            title_frame,
            text="📊 Export to Excel",
            font=ctk.CTkFont(size=14, weight="bold"),
            width=150,
            command=self.export_reports_to_excel
        )
        export_button.pack(side="right", padx=10)
        
        # Reports container
        reports_frame = ctk.CTkFrame(self.content_frame)
        reports_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Sales Report
        sales_frame = ctk.CTkFrame(reports_frame)
        sales_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(
            sales_frame,
            text="📊 Sales Report",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(anchor="w", padx=10, pady=10)
        
        try:
            connection = self.get_db_connection()
            if connection:
                cursor = connection.cursor(dictionary=True)
                
                # Total Sales
                cursor.execute("SELECT COUNT(*), SUM(total_amount) FROM orders WHERE status != 'cancelled'")
                sales_data = cursor.fetchone()
                total_orders = sales_data['COUNT(*)'] or 0
                total_revenue = sales_data['SUM(total_amount)'] or 0
                
                ctk.CTkLabel(
                    sales_frame,
                    text=f"Total Orders: {total_orders} | Total Revenue: ${total_revenue:.2f}",
                    font=ctk.CTkFont(size=14)
                ).pack(anchor="w", padx=20, pady=5)
                
                # Top Products
                cursor.execute("""
                    SELECT p.name, SUM(oi.quantity) as total_sold, SUM(oi.quantity * oi.price) as revenue
                    FROM order_items oi
                    JOIN products p ON oi.product_id = p.id
                    GROUP BY p.id
                    ORDER BY total_sold DESC
                    LIMIT 5
                """)
                top_products = cursor.fetchall()
                
                ctk.CTkLabel(
                    sales_frame,
                    text="Top 5 Products:",
                    font=ctk.CTkFont(size=14, weight="bold")
                ).pack(anchor="w", padx=20, pady=(10, 5))
                
                for i, product in enumerate(top_products, 1):
                    ctk.CTkLabel(
                        sales_frame,
                        text=f"  {i}. {product['name']} - {product['total_sold']} sold (${product['revenue']:.2f})",
                        font=ctk.CTkFont(size=12)
                    ).pack(anchor="w", padx=30, pady=2)
                
                # Inventory Report
                inventory_frame = ctk.CTkFrame(reports_frame)
                inventory_frame.pack(fill="x", padx=10, pady=10)
                
                ctk.CTkLabel(
                    inventory_frame,
                    text="📦 Inventory Report",
                    font=ctk.CTkFont(size=18, weight="bold")
                ).pack(anchor="w", padx=10, pady=10)
                
                cursor.execute("SELECT COUNT(*), SUM(quantity) FROM products")
                inv_data = cursor.fetchone()
                total_products = inv_data['COUNT(*)'] or 0
                total_stock = inv_data['SUM(quantity)'] or 0
                
                cursor.execute("SELECT COUNT(*) FROM products WHERE quantity <= reorder_level")
                low_stock_count = cursor.fetchone()['COUNT(*)']
                
                ctk.CTkLabel(
                    inventory_frame,
                    text=f"Total Products: {total_products} | Total Stock Units: {total_stock} | Low Stock Items: {low_stock_count}",
                    font=ctk.CTkFont(size=14)
                ).pack(anchor="w", padx=20, pady=5)
                
                # Customer Report
                customer_frame = ctk.CTkFrame(reports_frame)
                customer_frame.pack(fill="x", padx=10, pady=10)
                
                ctk.CTkLabel(
                    customer_frame,
                    text="👥 Customer Report",
                    font=ctk.CTkFont(size=18, weight="bold")
                ).pack(anchor="w", padx=10, pady=10)
                
                cursor.execute("SELECT COUNT(*) FROM users WHERE user_type = 'customer'")
                total_customers = cursor.fetchone()['COUNT(*)']
                
                # Top Customers
                cursor.execute("""
                    SELECT CONCAT(u.first_name, ' ', u.last_name) as full_name, u.email, COUNT(o.id) as order_count, SUM(o.total_amount) as total_spent
                    FROM users u
                    JOIN orders o ON u.id = o.customer_id
                    WHERE u.user_type = 'customer' AND o.status != 'cancelled'
                    GROUP BY u.id
                    ORDER BY total_spent DESC
                    LIMIT 5
                """)
                top_customers = cursor.fetchall()
                
                ctk.CTkLabel(
                    customer_frame,
                    text=f"Total Customers: {total_customers}",
                    font=ctk.CTkFont(size=14)
                ).pack(anchor="w", padx=20, pady=5)
                
                ctk.CTkLabel(
                    customer_frame,
                    text="Top 5 Customers by Spending:",
                    font=ctk.CTkFont(size=14, weight="bold")
                ).pack(anchor="w", padx=20, pady=(10, 5))
                
                for i, customer in enumerate(top_customers, 1):
                    ctk.CTkLabel(
                        customer_frame,
                        text=f"  {i}. {customer['full_name']} - {customer['order_count']} orders (${customer['total_spent']:.2f})",
                        font=ctk.CTkFont(size=12)
                    ).pack(anchor="w", padx=30, pady=2)
                
                cursor.close()
                
        except Error as e:
            messagebox.showerror("Error", f"Failed to load reports: {e}")
    
    def export_reports_to_excel(self):
        """Export all reports to an Excel file"""
        try:
            # Ask user where to save the file
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Reports as Excel"
            )
            
            if not filename:
                return  # User cancelled
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            connection = self.get_db_connection()
            if not connection:
                messagebox.showerror("Error", "Failed to connect to database")
                return
            
            cursor = connection.cursor(dictionary=True)
            
            # Get current user info for metadata
            current_user_name = f"{self.main_app.current_user.get('first_name', '')} {self.main_app.current_user.get('last_name', '')}".strip()
            current_user_email = self.main_app.current_user.get('email', 'N/A')
            access_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # ========== SALES REPORT ==========
            sales_sheet = wb.create_sheet("Sales Report")
            
            # Header with metadata
            sales_sheet['A1'] = "Smart Inventory Management System - Sales Report"
            sales_sheet['A1'].font = Font(size=16, bold=True)
            sales_sheet.merge_cells('A1:D1')
            
            sales_sheet['A2'] = f"Generated on: {access_date}"
            sales_sheet['A3'] = f"Accessed by: {current_user_name} ({current_user_email})"
            sales_sheet['A4'] = ""  # Empty row
            
            row = 5
            
            # Total Sales
            cursor.execute("SELECT COUNT(*) as total_orders, SUM(total_amount) as total_revenue FROM orders WHERE status != 'cancelled'")
            sales_data = cursor.fetchone()
            total_orders = sales_data['total_orders'] or 0
            total_revenue = sales_data['total_revenue'] or 0
            
            sales_sheet['A5'] = "Summary"
            sales_sheet['A5'].font = Font(size=14, bold=True)
            row = 6
            sales_sheet[f'A{row}'] = "Total Orders:"
            sales_sheet[f'B{row}'] = total_orders
            row += 1
            sales_sheet[f'A{row}'] = "Total Revenue:"
            sales_sheet[f'B{row}'] = f"${total_revenue:.2f}"
            row += 2
            
            # Top Products
            sales_sheet[f'A{row}'] = "Top 5 Products"
            sales_sheet[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            headers = ["Rank", "Product Name", "Total Sold", "Revenue"]
            for col, header in enumerate(headers, 1):
                cell = sales_sheet.cell(row, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            
            row += 1
            
            cursor.execute("""
                SELECT p.name, SUM(oi.quantity) as total_sold, SUM(oi.quantity * oi.price) as revenue
                FROM order_items oi
                JOIN products p ON oi.product_id = p.id
                GROUP BY p.id
                ORDER BY total_sold DESC
                LIMIT 5
            """)
            top_products = cursor.fetchall()
            
            for rank, product in enumerate(top_products, 1):
                sales_sheet.cell(row, 1, rank)
                sales_sheet.cell(row, 2, product['name'])
                sales_sheet.cell(row, 3, product['total_sold'])
                sales_sheet.cell(row, 4, f"${product['revenue']:.2f}")
                row += 1
            
            # Adjust column widths
            sales_sheet.column_dimensions['A'].width = 15
            sales_sheet.column_dimensions['B'].width = 30
            sales_sheet.column_dimensions['C'].width = 15
            sales_sheet.column_dimensions['D'].width = 15
            
            # ========== INVENTORY REPORT ==========
            inventory_sheet = wb.create_sheet("Inventory Report")
            
            inventory_sheet['A1'] = "Smart Inventory Management System - Inventory Report"
            inventory_sheet['A1'].font = Font(size=16, bold=True)
            inventory_sheet.merge_cells('A1:D1')
            
            inventory_sheet['A2'] = f"Generated on: {access_date}"
            inventory_sheet['A3'] = f"Accessed by: {current_user_name} ({current_user_email})"
            inventory_sheet['A4'] = ""
            
            row = 5
            
            cursor.execute("SELECT COUNT(*) as total_products, SUM(quantity) as total_stock FROM products")
            inv_data = cursor.fetchone()
            total_products = inv_data['total_products'] or 0
            total_stock = inv_data['total_stock'] or 0
            
            cursor.execute("SELECT COUNT(*) as low_stock_count FROM products WHERE quantity <= reorder_level")
            low_stock_data = cursor.fetchone()
            low_stock_count = low_stock_data['low_stock_count'] or 0
            
            inventory_sheet['A5'] = "Summary"
            inventory_sheet['A5'].font = Font(size=14, bold=True)
            row = 6
            inventory_sheet[f'A{row}'] = "Total Products:"
            inventory_sheet[f'B{row}'] = total_products
            row += 1
            inventory_sheet[f'A{row}'] = "Total Stock Units:"
            inventory_sheet[f'B{row}'] = total_stock
            row += 1
            inventory_sheet[f'A{row}'] = "Low Stock Items:"
            inventory_sheet[f'B{row}'] = low_stock_count
            row += 2
            
            # Low Stock Items Detail
            inventory_sheet[f'A{row}'] = "Low Stock Items (Qty <= Reorder Level)"
            inventory_sheet[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            headers = ["Product ID", "Product Name", "Current Quantity", "Reorder Level", "Status"]
            for col, header in enumerate(headers, 1):
                cell = inventory_sheet.cell(row, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            
            row += 1
            
            cursor.execute("""
                SELECT id, name, quantity, reorder_level
                FROM products
                WHERE quantity <= reorder_level
                ORDER BY quantity ASC
            """)
            low_stock_items = cursor.fetchall()
            
            for item in low_stock_items:
                inventory_sheet.cell(row, 1, item['id'])
                inventory_sheet.cell(row, 2, item['name'])
                inventory_sheet.cell(row, 3, item['quantity'])
                inventory_sheet.cell(row, 4, item['reorder_level'])
                inventory_sheet.cell(row, 5, "LOW STOCK")
                row += 1
            
            # Adjust column widths
            inventory_sheet.column_dimensions['A'].width = 12
            inventory_sheet.column_dimensions['B'].width = 30
            inventory_sheet.column_dimensions['C'].width = 18
            inventory_sheet.column_dimensions['D'].width = 15
            inventory_sheet.column_dimensions['E'].width = 15
            
            # ========== CUSTOMER REPORT ==========
            customer_sheet = wb.create_sheet("Customer Report")
            
            customer_sheet['A1'] = "Smart Inventory Management System - Customer Report"
            customer_sheet['A1'].font = Font(size=16, bold=True)
            customer_sheet.merge_cells('A1:E1')
            
            customer_sheet['A2'] = f"Generated on: {access_date}"
            customer_sheet['A3'] = f"Accessed by: {current_user_name} ({current_user_email})"
            customer_sheet['A4'] = ""
            
            row = 5
            
            cursor.execute("SELECT COUNT(*) as total_customers FROM users WHERE user_type = 'customer'")
            customer_data = cursor.fetchone()
            total_customers = customer_data['total_customers'] or 0
            
            customer_sheet['A5'] = "Summary"
            customer_sheet['A5'].font = Font(size=14, bold=True)
            row = 6
            customer_sheet[f'A{row}'] = "Total Customers:"
            customer_sheet[f'B{row}'] = total_customers
            row += 2
            
            # Top Customers
            customer_sheet[f'A{row}'] = "Top 5 Customers by Spending"
            customer_sheet[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            headers = ["Rank", "Customer Name", "Email", "Order Count", "Total Spent"]
            for col, header in enumerate(headers, 1):
                cell = customer_sheet.cell(row, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            
            row += 1
            
            cursor.execute("""
                SELECT CONCAT(u.first_name, ' ', u.last_name) as full_name, u.email, COUNT(o.id) as order_count, SUM(o.total_amount) as total_spent
                FROM users u
                JOIN orders o ON u.id = o.customer_id
                WHERE u.user_type = 'customer' AND o.status != 'cancelled'
                GROUP BY u.id
                ORDER BY total_spent DESC
                LIMIT 5
            """)
            top_customers = cursor.fetchall()
            
            for rank, customer in enumerate(top_customers, 1):
                customer_sheet.cell(row, 1, rank)
                customer_sheet.cell(row, 2, customer['full_name'])
                customer_sheet.cell(row, 3, customer['email'])
                customer_sheet.cell(row, 4, customer['order_count'])
                customer_sheet.cell(row, 5, f"${customer['total_spent']:.2f}")
                row += 1
            
            # Adjust column widths
            customer_sheet.column_dimensions['A'].width = 8
            customer_sheet.column_dimensions['B'].width = 25
            customer_sheet.column_dimensions['C'].width = 30
            customer_sheet.column_dimensions['D'].width = 15
            customer_sheet.column_dimensions['E'].width = 15
            
            cursor.close()
            
            # Save the workbook
            wb.save(filename)
            
            messagebox.showinfo("Success", f"Reports exported successfully to:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export reports: {str(e)}")
    
    def logout(self):
        if messagebox.askyesno("Logout", "Are you sure you want to logout?"):
            self.main_app.logout()


class AddSupplierDialog(ctk.CTkToplevel):
    def __init__(self, main_app, admin_dashboard):
        super().__init__(main_app)
        self.main_app = main_app
        self.admin_dashboard = admin_dashboard
        
        self.title("Add New Supplier")
        self.geometry("500x500")
        self.resizable(False, False)
        
        # Center the window
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
        # Create form
        form_frame = ctk.CTkFrame(self)
        form_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        ctk.CTkLabel(
            form_frame,
            text="Add New Supplier",
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(pady=(10, 20))
        
        # Supplier Name
        ctk.CTkLabel(form_frame, text="Supplier Name:", font=ctk.CTkFont(size=14)).pack(anchor="w", padx=20, pady=(10, 5))
        self.name_entry = ctk.CTkEntry(form_frame, width=400, placeholder_text="Enter supplier name")
        self.name_entry.pack(padx=20, pady=(0, 10))
        
        # Contact Person
        ctk.CTkLabel(form_frame, text="Contact Person:", font=ctk.CTkFont(size=14)).pack(anchor="w", padx=20, pady=(10, 5))
        self.contact_entry = ctk.CTkEntry(form_frame, width=400, placeholder_text="Enter contact person name")
        self.contact_entry.pack(padx=20, pady=(0, 10))
        
        # Email
        ctk.CTkLabel(form_frame, text="Email:", font=ctk.CTkFont(size=14)).pack(anchor="w", padx=20, pady=(10, 5))
        self.email_entry = ctk.CTkEntry(form_frame, width=400, placeholder_text="Enter email address")
        self.email_entry.pack(padx=20, pady=(0, 10))
        
        # Phone
        ctk.CTkLabel(form_frame, text="Phone:", font=ctk.CTkFont(size=14)).pack(anchor="w", padx=20, pady=(10, 5))
        self.phone_entry = ctk.CTkEntry(form_frame, width=400, placeholder_text="Enter phone number")
        self.phone_entry.pack(padx=20, pady=(0, 10))
        
        # Address
        ctk.CTkLabel(form_frame, text="Address:", font=ctk.CTkFont(size=14)).pack(anchor="w", padx=20, pady=(10, 5))
        self.address_entry = ctk.CTkEntry(form_frame, width=400, placeholder_text="Enter supplier address")
        self.address_entry.pack(padx=20, pady=(0, 10))
        
        # Buttons
        button_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        button_frame.pack(pady=20)
        
        ctk.CTkButton(
            button_frame,
            text="Add Supplier",
            width=150,
            command=self.add_supplier
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            button_frame,
            text="Cancel",
            width=150,
            fg_color="gray",
            command=self.destroy
        ).pack(side="left", padx=10)
    
    def add_supplier(self):
        name = self.name_entry.get().strip()
        contact_person = self.contact_entry.get().strip()
        email = self.email_entry.get().strip()
        phone = self.phone_entry.get().strip()
        address = self.address_entry.get().strip()
        
        # Form validation
        if not name:
            messagebox.showerror("Error", "Please enter supplier name")
            return
        
        # Validate email if provided
        if email:
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, email):
                messagebox.showerror("Error", "Please enter a valid email address")
                return
        
        # Validate phone if provided
        if phone:
            phone_cleaned = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')
            if not phone_cleaned.isdigit() or len(phone_cleaned) < 10:
                messagebox.showerror("Error", "Please enter a valid phone number (at least 10 digits)")
                return
        
        try:
            connection = db_config.create_connection()
            if connection:
                cursor = connection.cursor()
                
                # Insert supplier
                cursor.execute(
                    """INSERT INTO suppliers (name, contact_person, email, phone, address)
                    VALUES (%s, %s, %s, %s, %s)""",
                    (name, contact_person or None, email or None, phone or None, address or None)
                )
                
                connection.commit()
                cursor.close()
                connection.close()
                
                messagebox.showinfo("Success", "Supplier added successfully!")
                self.destroy()
                # Refresh supplier list after dialog closes
                if hasattr(self.admin_dashboard, 'load_suppliers'):
                    self.main_app.after(100, self.admin_dashboard.load_suppliers)
                
        except Error as e:
            messagebox.showerror("Error", f"Failed to add supplier: {e}")


class EditSupplierDialog(ctk.CTkToplevel):
    def __init__(self, main_app, admin_dashboard, supplier_id):
        super().__init__(main_app)
        self.main_app = main_app
        self.admin_dashboard = admin_dashboard
        self.supplier_id = supplier_id
        
        self.title("Edit Supplier")
        self.geometry("500x500")
        self.resizable(False, False)
        
        # Center the window
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
        # Create form
        form_frame = ctk.CTkFrame(self)
        form_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        ctk.CTkLabel(
            form_frame,
            text="Edit Supplier",
            font=ctk.CTkFont(size=20, weight="bold")
        ).pack(pady=(10, 20))
        
        # Supplier Name
        ctk.CTkLabel(form_frame, text="Supplier Name:", font=ctk.CTkFont(size=14)).pack(anchor="w", padx=20, pady=(10, 5))
        self.name_entry = ctk.CTkEntry(form_frame, width=400)
        self.name_entry.pack(padx=20, pady=(0, 10))
        
        # Contact Person
        ctk.CTkLabel(form_frame, text="Contact Person:", font=ctk.CTkFont(size=14)).pack(anchor="w", padx=20, pady=(10, 5))
        self.contact_entry = ctk.CTkEntry(form_frame, width=400)
        self.contact_entry.pack(padx=20, pady=(0, 10))
        
        # Email
        ctk.CTkLabel(form_frame, text="Email:", font=ctk.CTkFont(size=14)).pack(anchor="w", padx=20, pady=(10, 5))
        self.email_entry = ctk.CTkEntry(form_frame, width=400)
        self.email_entry.pack(padx=20, pady=(0, 10))
        
        # Phone
        ctk.CTkLabel(form_frame, text="Phone:", font=ctk.CTkFont(size=14)).pack(anchor="w", padx=20, pady=(10, 5))
        self.phone_entry = ctk.CTkEntry(form_frame, width=400)
        self.phone_entry.pack(padx=20, pady=(0, 10))
        
        # Address
        ctk.CTkLabel(form_frame, text="Address:", font=ctk.CTkFont(size=14)).pack(anchor="w", padx=20, pady=(10, 5))
        self.address_entry = ctk.CTkEntry(form_frame, width=400)
        self.address_entry.pack(padx=20, pady=(0, 10))
        
        # Buttons
        button_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        button_frame.pack(pady=20)
        
        ctk.CTkButton(
            button_frame,
            text="Update Supplier",
            width=150,
            command=self.update_supplier
        ).pack(side="left", padx=10)
        
        ctk.CTkButton(
            button_frame,
            text="Cancel",
            width=150,
            fg_color="gray",
            command=self.destroy
        ).pack(side="left", padx=10)
        
        # Load supplier data
        self.load_supplier_data()
    
    def load_supplier_data(self):
        try:
            connection = db_config.create_connection()
            if connection:
                cursor = connection.cursor(dictionary=True)
                cursor.execute("SELECT * FROM suppliers WHERE id = %s", (self.supplier_id,))
                supplier = cursor.fetchone()
                
                if supplier:
                    self.name_entry.insert(0, supplier['name'])
                    self.contact_entry.insert(0, supplier['contact_person'] or '')
                    self.email_entry.insert(0, supplier['email'] or '')
                    self.phone_entry.insert(0, supplier['phone'] or '')
                    self.address_entry.insert(0, supplier['address'] or '')
                
                cursor.close()
                connection.close()
                
        except Error as e:
            messagebox.showerror("Error", f"Failed to load supplier data: {e}")
    
    def update_supplier(self):
        name = self.name_entry.get().strip()
        contact_person = self.contact_entry.get().strip()
        email = self.email_entry.get().strip()
        phone = self.phone_entry.get().strip()
        address = self.address_entry.get().strip()
        
        # Form validation
        if not name:
            messagebox.showerror("Error", "Please enter supplier name")
            return
        
        # Validate email if provided
        if email:
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, email):
                messagebox.showerror("Error", "Please enter a valid email address")
                return
        
        # Validate phone if provided
        if phone:
            phone_cleaned = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')
            if not phone_cleaned.isdigit() or len(phone_cleaned) < 10:
                messagebox.showerror("Error", "Please enter a valid phone number (at least 10 digits)")
                return
        
        try:
            connection = db_config.create_connection()
            if connection:
                cursor = connection.cursor()
                
                # Update supplier
                cursor.execute(
                    """UPDATE suppliers 
                    SET name = %s, contact_person = %s, email = %s, phone = %s, address = %s
                    WHERE id = %s""",
                    (name, contact_person or None, email or None, phone or None, address or None, self.supplier_id)
                )
                
                connection.commit()
                cursor.close()
                connection.close()
                
                messagebox.showinfo("Success", "Supplier updated successfully!")
                self.destroy()
                # Refresh supplier list after dialog closes
                if hasattr(self.admin_dashboard, 'load_suppliers'):
                    self.main_app.after(100, self.admin_dashboard.load_suppliers)
                
        except Error as e:
            messagebox.showerror("Error", f"Failed to update supplier: {e}")
